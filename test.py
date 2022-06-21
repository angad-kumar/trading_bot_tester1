import lib as lb
import openpyxl 
import time
from datetime import datetime
from openpyxl import load_workbook
# load excel with its path
wrkbk = load_workbook('bank_nifty_option_chain.xlsx')
# to get the active work sheet
sh = wrkbk.active

 
# Print value of cell object
# using the value attribute
# print(cell_obj.value)

# sheet['A1'] = "time"
# sheet['B1'] = "price"
# sheet['C1'] = "pcr"
# sheet['D1'] = "change_oi_pcr"

while(1):
    row = sh.max_row + 1
    now = datetime.now()
    current_time = now.strftime("%H:%M")
    current_time_in_int = now.hour * 60 + now.minute
    
    if current_time_in_int > 0:
        # print(current_time," :", lb.current_pcr_value())
        sh.cell(row=row, column=1).value = current_time
        sh.cell(row=row, column=2).value = lb.current_value()
        sh.cell(row=row, column=3).value = lb.current_pcr_value()
        sh.cell(row=row, column=4).value = lb.change_in_oi_current_pcr_value()
        # row += 1
        wrkbk.save('bank_nigty_option_chain.xlsx')
        print(row)
        time.sleep(200)
    
 
# while(1):
#     print(lb.current_pcr_value())  
#     print(lb.change_in_oi_current_pcr_value())
#     time.sleep(10)


# print(lb.change_in_oi_current_pcr_value())
# print(lb.current_pcr_value())